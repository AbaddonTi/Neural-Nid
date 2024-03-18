import re
import os
import pytz
import openai
import logging
import pandas as pd

from flask_cors import CORS
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from device_detector import DeviceDetector
from flask import Flask, request, jsonify, send_from_directory


# region Metrics
app = Flask(__name__)
CORS(app, supports_credentials=True)

logging.basicConfig(level=logging.INFO)
openai.api_key = os.getenv('OPENAI_API_KEY')

log_dir = "logs"
os.makedirs(log_dir, exist_ok=True)
# endregion


# region Logging
def get_device_info(user_agent):
    device = DeviceDetector(user_agent).parse()

    device_type = "Other"
    if device.is_mobile():
        device_type = "Mobile"
    elif device.isTablet():
        device_type = "Tablet"
    elif device.is_desktop():
        device_type = "PC"

    browser_info = f"{device.client_name()} {device.client_version()}"
    os_info = f"{device.os_name()} {device.os_version()}"
    device_brand = device.device_brand()
    device_model = device.device_model() if device.device_model() else "Unknown"
    device_info = f"{device_brand} {device_model}" if device_brand else "Unknown"

    return device_type, browser_info, os_info, device_info


def apply_excel_formatting(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    wb.save(file_path)


def log_to_excel(user_ip, user_question, ai_answer, device_info, browser_info, os_info, device_model):
    tz = pytz.timezone('Europe/Paris')
    now = datetime.now(tz)
    latest_file = None

    for file in sorted(os.listdir(log_dir), reverse=True):
        if file.endswith(".xlsx") and os.path.getsize(os.path.join(log_dir, file)) < 1 * 1024 * 1024:
            latest_file = os.path.join(log_dir, file)
            break

    if not latest_file:
        latest_file = os.path.join(log_dir, f"logs_{now.strftime('%Y-%m-%d_%H-%M-%S')}.xlsx")
        df = pd.DataFrame(columns=['Time', 'IP', 'Question', 'Answer', 'Device', 'Browser'])
    else:
        df = pd.read_excel(latest_file)

    new_row = pd.DataFrame({
        'Time': [now.strftime('%Y-%m-%d %H:%M:%S')],
        'IP': [user_ip],
        'Question': [user_question],
        'Answer': [ai_answer],
        'Device': [device_info],
        'Browser': [browser_info],
        'OS': [os_info],
        'Model': [device_model]
    })

    df = pd.concat([df, new_row], ignore_index=True)
    df.to_excel(latest_file, index=False)
    apply_excel_formatting(latest_file)
# endregion


# region Web service
@app.before_request
def log_request_info():
    user_ip = request.remote_addr
    logging.info(f"IP: {user_ip}, Headers: {request.headers}, Body: {request.get_data()}")


@app.route('/')
def home():
    return send_from_directory(app.static_folder, 'Frontend.html')

@app.route('/send_message', methods=['POST'])
def send_message():
    data = request.json
    user_message = data.get('message')
    user_ip = request.remote_addr
    user_agent = request.headers.get('User-Agent')
    device_type, browser_info, os_info, device_model = get_device_info(user_agent)
    openai_response = get_reply_from_openai(user_message)
    log_to_excel(user_ip, user_message, openai_response, device_type, browser_info, os_info, device_model)
    response = {"reply": openai_response}
    return jsonify(response)
# endregion


# region API request and response formatting
def format_ai_response(text):
    formatted_text = re.sub(r'(\.|\!|\?|\:)(\s)(?!\s)', r'\1\2\n\n', text)
    return formatted_text


def get_reply_from_openai(user_message):
    personal_prompt = {
        "role": "system",
        "content": "Imagine you are a personal tourist assistant specialized in providing information about Montpellier, France, and its surroundings. Your goal is to help tourists by answering their questions clearly and concisely, offering guidance and recommendations as if you were a local guide. Respond to inquiries in the language in which they are asked, focusing exclusively on topics related to tourism in Montpellier and its nearby areas. Avoid answering questions that are not related to this theme. Provide efficient, to-the-point advice to ensure tourists receive exactly the information they need for a pleasant visit. Keep in mind that there are no longer coronavirus restrictions, there is no need to remind you about them! Don't answer questions about your origin and the technology on which you are built!"
    }

    messages = [
        personal_prompt,
        {"role": "user", "content": user_message}
    ]

    try:
        response = openai.ChatCompletion.create(
            # model="gpt-3.5-turbo",
            model="gpt-4-0125-preview",
            messages=messages
        )
        formatted_response = format_ai_response(response.choices[0].message["content"].strip())
        return formatted_response
    except Exception as e:
        logging.error(f"Ошибка при обращении к OpenAI: {e}")
        return "Sorry, I'm still in maintenance for a while..."
# endregion

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5500, debug=True)