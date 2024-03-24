import os
import pandas as pd
import pytz
from flask import send_file


from flask import Flask, request, jsonify
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment


app = Flask(__name__)


log_dir = "logs"
os.makedirs(log_dir, exist_ok=True)


def apply_excel_formatting(file_path):
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        wb.save(file_path)
    except Exception as e:
        print(f"Error applying Excel formatting: {e}")


@app.route('/log', methods=['POST'])
def log_to_csv():
    try:
        data = request.json

        tz = pytz.timezone('Europe/Paris')
        now = datetime.now(tz)
        latest_file_csv = os.path.join(log_dir, f"logs.csv")

        if not os.path.exists(latest_file_csv):
            df = pd.DataFrame(columns=['Time', 'IP', 'Question', 'Answer', 'Device', 'Browser', 'OS'])
        else:
            df = pd.read_csv(latest_file_csv)

        new_row = {
            'Time': now.strftime('%Y-%m-%d %H:%M:%S'),
            'IP': data.get('IP'),
            'Question': data.get('Question'),
            'Answer': data.get('Answer'),
            'Device': data.get('Device'),
            'Browser': data.get('Browser'),
            'OS': data.get('OS'),
        }

        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
        df.to_csv(latest_file_csv, index=False)

        return jsonify({"message": "Logged successfully"}), 200
    except Exception as e:
        return jsonify({"error": f"Failed to log data: {str(e)}"}), 500



@app.route('/get_latest_log', methods=['GET'])
def get_latest_log():
    try:
        latest_file_csv = os.path.join(log_dir, "logs.csv")
        if os.path.exists(latest_file_csv):
            return send_file(latest_file_csv, as_attachment=True)
        else:
            return jsonify({"message": "No logs found"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500



if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5600)
